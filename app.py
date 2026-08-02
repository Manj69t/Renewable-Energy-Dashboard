import streamlit as st
import pandas as pd
import numpy as np
import os
import io
from openpyxl.utils import column_index_from_string, get_column_letter

# ==========================================
# GLOBAL SETUP & HELPER FUNCTIONS
# ==========================================
# Set page config MUST be the first Streamlit command
st.set_page_config(page_title="Integrated RTC & NHPC Power Models", page_icon="⚡", layout="wide")

@st.cache_data
def load_data_nhpc(uploaded_file, sheet, skip_rows):
    """Helper function exclusively for the NHPC 25-Year Model using uploaded file"""
    return pd.read_excel(uploaded_file, sheet_name=sheet, skiprows=skip_rows - 1, header=None, engine="openpyxl")


# =========================================================================================
# SUBLAYER 1: PPA RTC Power Model V3 (Seasonal Compliance)
# =========================================================================================
def run_model_1_rtc_v3():
    st.title("⚡ PPA RTC Power Model (Seasonal Compliance & DFR)")

    st.sidebar.header("📁 Raw Data Upload")
    # CHANGED: Replaced local text path with a file uploader
    uploaded_file = st.sidebar.file_uploader("Upload Raw Data Excel File (.xlsx, .xlsm)", type=["xlsx", "xlsm"], key="m1_file")
    sheet_name = st.sidebar.text_input("Sheet Name", value="Raw Data", key="m1_sheet")

    wind_col_letter = st.sidebar.text_input("📊 WIND Per-Unit (PU) Column Letter", value="C", key="m1_wcol")
    solar_col_letter = st.sidebar.text_input("☀️ SOLAR Per-Unit (PU) Column Letter", value="D", key="m1_scol")
    data_start_row = st.sidebar.number_input("🔢 Starting Row for Raw Data (1-based)", value=6, min_value=1, step=1, key="m1_row")

    st.sidebar.markdown("---")
    st.sidebar.header("⏰ Peak Hours Configuration")
    enable_peak_hours = st.sidebar.checkbox("Enable Peak/Non-Peak Brackets", value=True, key="m1_peak_cb")

    if enable_peak_hours:
        morn_peak_start, morn_peak_end = st.sidebar.slider("Morning Peak Window", 0.0, 24.0, (0.0, 0.0), step=0.25, format="%.2f Hours", key="m1_mpeak")
        eve_peak_start, eve_peak_end = st.sidebar.slider("Evening Peak Window", 0.0, 24.0, (19.0, 21.0), step=0.25, format="%.2f Hours", key="m1_epeak")
    else:
        st.sidebar.info("ℹ️ Peak hours disabled. Operational spectrum locked to All Non-Peak (0).")

    st.sidebar.markdown("---")
    st.sidebar.header("📊 Scenario & Probability Controls")
    wind_probability_scenario = st.sidebar.selectbox("Wind Probability Scenario (C34)", options=["P50", "P75", "P90"], index=0, key="m1_wps")
    solar_probability_scenario = st.sidebar.selectbox("Solar Probability Scenario (C38)", options=["P50", "P75", "P90"], index=0, key="m1_sps")
    p90_factor = st.sidebar.number_input("P90 Factor Value (C33)", value=0.924, step=0.001, format="%.3f", key="m1_p90")
    all_loss_factor = st.sidebar.number_input("Wind All Loss Factor (C43)", value=-0.0145, step=0.05, format="%.2f", key="m1_loss")

    st.sidebar.markdown("---")
    st.sidebar.header("🎯 Seasonal Compliance Factors")
    winter_compliance_factor = st.sidebar.number_input("Winter Compliance Factor (C16)", value=0.55, min_value=0.01, max_value=1.00, step=0.01, format="%.2f", key="m1_win")
    summer_compliance_factor = st.sidebar.number_input("Summer Compliance Factor (C17)", value=0.55, min_value=0.01, max_value=1.00, step=0.01, format="%.2f", key="m1_sum")

    st.sidebar.markdown("---")
    st.sidebar.header("⚙️ Dynamic Plant Specifications")
    grid_connectivity = st.sidebar.number_input("Grid Connectivity Limit (MW - C1)", value=120.0, step=5.0, format="%.1f", key="m1_gc")
    exchange_connectivity = st.sidebar.number_input("Exchange Connectivity Limit (MW - C2)", value=180.0, step=5.0, format="%.1f", key="m1_ec")
    solar_capacity = st.sidebar.number_input("Solar Capacity (MW)", value=154.0, step=1.0, format="%.1f", key="m1_scap")
    wind_capacity = st.sidebar.number_input("Wind Capacity (MW)", value=225.0, step=1.0, format="%.1f", key="m1_wcap")
    pcs_capacity = st.sidebar.number_input("PCS Total Capacity (MW - C11)", value=100.0, step=1.0, format="%.1f", key="m1_pcs")
    bess_installed_capacity = st.sidebar.number_input("BESS Installed Capacity (MWh - C12)", value=600.0, step=1.0, format="%.1f", key="m1_bic")
    bess_usable_capacity = st.sidebar.number_input("BESS Usable Capacity (MWh - C14)", value=586.1, step=1.0, format="%.1f", key="m1_buc")

    st.sidebar.markdown("---")
    st.sidebar.header("📉 Variable Loss & Efficiency Factors")
    solar_hv_loss_factor = st.sidebar.number_input("Solar HV Loss Factor (D22)", value=0.9670, step=0.001, format="%.4f", key="m1_shv")
    wind_hv_loss_factor = st.sidebar.number_input("Wind HV Loss Factor (E22)", value=0.9718, step=0.001, format="%.4f", key="m1_whv")
    ac_charging_eff = st.sidebar.number_input("AC Charging Efficiency (C22)", value=0.936, step=0.001, format="%.3f", key="m1_acc")
    total_charging_eff = st.sidebar.number_input("Total Charging Efficiency (C24)", value=0.912, step=0.001, format="%.3f", key="m1_tce")
    ac_discharging_eff_pcs = st.sidebar.number_input("AC Discharging Eff at PCS Terminal (C23)", value=0.936, step=0.001, format="%.3f", key="m1_acd")
    total_discharging_eff = st.sidebar.number_input("Total Discharging Efficiency (C25)", value=0.917, step=0.001, format="%.3f", key="m1_tde")
    ac_discharging_eff_poc = st.sidebar.number_input("AC Discharging Eff till POC (C27)", value=0.886, step=0.001, format="%.3f", key="m1_acdpoc")
    solar_degradation_factor = st.sidebar.number_input("Solar Degradation Factor (C41)", value=1.000, step=0.005, format="%.3f", key="m1_sdeg")
    wind_degradation_factor = st.sidebar.number_input("Wind Degradation Factor (C40)", value=1.000, step=0.005, format="%.3f", key="m1_wdeg")

    wind_idx = column_index_from_string(wind_col_letter) - 1
    solar_idx = column_index_from_string(solar_col_letter) - 1

    st.markdown("### 🗺️ Raw Input Sheet Mapping Guidelines")
    mapping_data = {
        "Excel Expected Profile Component": ["WIND Generation Per-Unit (PU) Data stream", "SOLAR Generation Per-Unit (PU) Data stream", "Data Starting Coordinate Point"],
        "Mapped Target Column Letter": [wind_col_letter.upper(), solar_col_letter.upper(), f"Row {data_start_row} (1-based)"],
        "Zero-Based Python Index Mapping": [f"Index {wind_idx}", f"Index {solar_idx}", f"Skipping first {data_start_row - 1} row(s)"]
    }
    st.table(pd.DataFrame(mapping_data))

    # CHANGED: Check if file is uploaded instead of local path existence
    if uploaded_file is not None:
        try:
            df_raw = pd.read_excel(uploaded_file, sheet_name=sheet_name, skiprows=int(data_start_row - 1), header=None, engine="openpyxl")
            df = pd.DataFrame()
            
            generated_timestamps = pd.date_range(start="1990-01-01 00:00", periods=len(df_raw), freq="15min")
            df["Date"] = generated_timestamps.strftime("%d-%m-%Y")
            df["Hour"] = generated_timestamps.hour
            df["Time"] = generated_timestamps.strftime("%H:%M")
            
            months = generated_timestamps.month
            hours = generated_timestamps.hour
            
            if enable_peak_hours:
                decimal_hours = generated_timestamps.hour + generated_timestamps.minute / 60.0
                is_morn_peak = (morn_peak_start != morn_peak_end) & (decimal_hours >= morn_peak_start) & (decimal_hours <= morn_peak_end)
                is_eve_peak = (eve_peak_start != eve_peak_end) & (decimal_hours >= eve_peak_start) & (decimal_hours <= eve_peak_end)
                df["Peak/Non Peak"] = np.where(is_morn_peak | is_eve_peak, 1, 0)
            else:
                df["Peak/Non Peak"] = 0
            
            df["DFR"] = np.where(df["Peak/Non Peak"] == 0, 0.70, 0.90)
            df["Peak Demand @ POC (MW)"] = grid_connectivity / 4.0
            
            is_winter_cond = ((months <= 3) | (months >= 10)) & (hours >= 7) & (hours < 17)
            is_summer_cond = ((months >= 4) & (months <= 9)) & (hours >= 6) & (hours < 18)
            
            df["Actual Peak Demand @ POC (MW)"] = np.where(
                is_winter_cond,
                df["Peak Demand @ POC (MW)"] * winter_compliance_factor,
                np.where(is_summer_cond, df["Peak Demand @ POC (MW)"] * summer_compliance_factor, df["Peak Demand @ POC (MW)"])
            )
            
            if wind_idx >= len(df_raw.columns) or solar_idx >= len(df_raw.columns):
                st.error(f"🚨 Configuration Error: Mapped Column Letters fall out of bounds!")
                st.stop()
                
            raw_wind_pu = pd.to_numeric(df_raw.iloc[:, wind_idx], errors='coerce').fillna(0.0)
            raw_solar_pu = pd.to_numeric(df_raw.iloc[:, solar_idx], errors='coerce').fillna(0.0)
            
            wind_multiplier = 1.0 if wind_probability_scenario in ["P50", "P75"] else p90_factor
            converted_wind_pu = (raw_wind_pu * wind_multiplier) * (1.0 + all_loss_factor)
            
            solar_multiplier = 1.0 if solar_probability_scenario in ["P50", "P75"] else p90_factor
            converted_solar_pu = raw_solar_pu * solar_multiplier
            
            df["WIND Power @ 33 kV"] = converted_wind_pu * (wind_capacity / 4.0) * wind_degradation_factor
            df["Solar Power @ 33 kV"] = converted_solar_pu * (solar_capacity / 4.0) * solar_degradation_factor
            df["Total Power @ POC"] = df["WIND Power @ 33 kV"] + df["Solar Power @ 33 kV"]
            
            df["Power Beyond PPA Demand"] = np.where(df["Total Power @ POC"] > df["Actual Peak Demand @ POC (MW)"], df["Total Power @ POC"] - df["Actual Peak Demand @ POC (MW)"], 0.0)
            df["Power Available for Charging POC"] = (df["Total Power @ POC"] * solar_hv_loss_factor) - df["Actual Peak Demand @ POC (MW)"]
            df["Is it Charge (C)/ Discharge (D)"] = np.where(df["Power Available for Charging POC"] < 0.0, "D", "C")
            
            df["If D, then Energy Discharges from PCS"] = np.where(df["Is it Charge (C)/ Discharge (D)"] == "D", df["Power Available for Charging POC"] / ac_discharging_eff_poc, 0.0)
            pcs_dis_limit_mwh = (pcs_capacity / ac_discharging_eff_pcs) / 4.0
            df["Discharging Power Limited to PCS Rating"] = np.where(df["If D, then Energy Discharges from PCS"].abs() > pcs_dis_limit_mwh, -pcs_dis_limit_mwh, df["If D, then Energy Discharges from PCS"])
            
            size = len(df)
            energy_goes_for_ess, charging_power_limited_to_ess, daily_charging_limit, soc_tracker = np.zeros(size), np.zeros(size), np.zeros(size), np.zeros(size)
            charge_discharge_increment, charging_energy, discharging_energy, abs_of_increment = np.zeros(size), np.zeros(size), np.zeros(size), np.zeros(size)
            power_beyond_full_charging, bess_standby_aux = np.zeros(size), np.zeros(size)
            
            current_soc = bess_usable_capacity
            dates, modes = df["Date"].values, df["Is it Charge (C)/ Discharge (D)"].values
            total_power_poc, power_beyond_ppa = df["Total Power @ POC"].values, df["Power Beyond PPA Demand"].values
            discharging_power_limited = df["Discharging Power Limited to PCS Rating"].values
            
            pcs_chg_limit_mwh = (pcs_capacity * ac_charging_eff) / 4.0
            max_daily_charging = bess_usable_capacity * 2.0
            
            for i in range(size):
                prev_daily_limit = 0.0 if i == 0 or dates[i] != dates[i-1] else daily_charging_limit[i-1]
                prev_soc = current_soc
                
                energy_goes_for_ess[i] = power_beyond_ppa[i] * total_charging_eff if (prev_soc < bess_usable_capacity) and (modes[i] == "C") else 0.0
                charging_power_limited_to_ess[i] = pcs_chg_limit_mwh if energy_goes_for_ess[i] > pcs_chg_limit_mwh else energy_goes_for_ess[i]
                daily_charging_limit[i] = min(prev_daily_limit + charging_power_limited_to_ess[i], max_daily_charging) if modes[i] == "C" else prev_daily_limit
                
                trial_soc = prev_soc + discharging_power_limited[i] + charging_power_limited_to_ess[i]
                if (modes[i] == "D") and (trial_soc < 0.0): current_soc = 0.0
                elif trial_soc > bess_usable_capacity: current_soc = bess_usable_capacity
                else: current_soc = prev_soc + discharging_power_limited[i] if (daily_charging_limit[i] >= max_daily_charging) else trial_soc
                        
                soc_tracker[i] = current_soc
                charge_discharge_increment[i] = current_soc - prev_soc
                charging_energy[i] = charge_discharge_increment[i] if charge_discharge_increment[i] > 0.0 else 0.0
                discharging_energy[i] = charge_discharge_increment[i] if charge_discharge_increment[i] < 0.0 else 0.0
                abs_of_increment[i] = abs(charge_discharge_increment[i])
                surplus_ac = total_power_poc[i] - (abs_of_increment[i] / ac_charging_eff)
                power_beyond_full_charging[i] = surplus_ac if surplus_ac > 0.0 else 0.0
                bess_standby_aux[i] = (3.5 * bess_installed_capacity / 4.0) / 1000.0 if charge_discharge_increment[i] == 0.0 else 0.0
                
            df["If C, Energy Goes for ESS"] = energy_goes_for_ess
            df["Charging Power Limited to ESS Capacity"] = charging_power_limited_to_ess
            df["Daily Charging Limit"] = daily_charging_limit
            df["SOC"] = soc_tracker
            df["Charge/Discharge Increment"] = charge_discharge_increment
            df["Charging Energy"] = charging_energy
            df["Discharging Energy"] = discharging_energy
            df["ABS of 12"] = abs_of_increment
            df["Power Beyond Full Charging"] = power_beyond_full_charging
            df["BESS Standby Aux"] = bess_standby_aux
            
            df["Power Available for Grid at 220 KV @ POC"] = np.where(
                df["Is it Charge (C)/ Discharge (D)"] == "D",
                df["Total Power @ POC"] - (df["Charge/Discharge Increment"] * total_discharging_eff),
                df["Power Beyond Full Charging"]
            ) * solar_hv_loss_factor - df["BESS Standby Aux"]
            
            connectivity_limit_mwh = grid_connectivity / 4.0
            df["Power Injected to SECI @ POC"] = np.where(df["Power Available for Grid at 220 KV @ POC"] > df["Actual Peak Demand @ POC (MW)"], df["Actual Peak Demand @ POC (MW)"], df["Power Available for Grid at 220 KV @ POC"])
            df["Excess Power beyond evacuation level @ POC"] = df["Power Available for Grid at 220 KV @ POC"] - df["Power Injected to SECI @ POC"]
            connectivity_exchange_limit_mwh = exchange_connectivity / 4.0
            df["Energy to Exchange @ POC"] = np.where(df["Excess Power beyond evacuation level @ POC"] < connectivity_exchange_limit_mwh, df["Excess Power beyond evacuation level @ POC"], connectivity_exchange_limit_mwh)
            df["Unsold Energy @ POC"] = df["Excess Power beyond evacuation level @ POC"] - df["Energy to Exchange @ POC"]
            df["Availability"] = df["Power Injected to SECI @ POC"] / connectivity_limit_mwh
            df["DFR2"] = np.where(df["Availability"] < df["DFR"], (connectivity_limit_mwh * df["DFR"]) - df["Power Available for Grid at 220 KV @ POC"], 0.0)
            
            # --- Seasonal Compliance Matrix ---
            m_peak = df["Peak/Non Peak"] == 1
            m_summer = (months >= 4) & (months <= 9)
            m_summer_solar = m_summer & (hours >= 6) & (hours < 18)
            m_summer_non_solar = m_summer & ((hours < 6) | (hours >= 18))
            m_winter = (months <= 3) | (months >= 10)
            m_winter_solar = m_winter & (hours >= 7) & (hours < 17)
            m_winter_non_solar = m_winter & ((hours < 7) | (hours >= 17))
            
            demand_peak_eve = df.loc[m_peak, "Peak Demand @ POC (MW)"].sum() / 1000.0
            demand_summer_solar = df.loc[m_summer_solar, "Peak Demand @ POC (MW)"].sum() / 1000.0
            demand_summer_non_solar = df.loc[m_summer_non_solar, "Peak Demand @ POC (MW)"].sum() / 1000.0
            demand_winter_solar = df.loc[m_winter_solar, "Peak Demand @ POC (MW)"].sum() / 1000.0
            demand_winter_non_solar = df.loc[m_winter_non_solar, "Peak Demand @ POC (MW)"].sum() / 1000.0
            
            supply_peak_eve = df.loc[m_peak, "Power Injected to SECI @ POC"].sum() / 1000.0
            supply_summer_solar = df.loc[m_summer_solar, "Power Injected to SECI @ POC"].sum() / 1000.0
            supply_summer_non_solar = df.loc[m_summer_non_solar, "Power Injected to SECI @ POC"].sum() / 1000.0
            supply_winter_solar = df.loc[m_winter_solar, "Power Injected to SECI @ POC"].sum() / 1000.0
            supply_winter_non_solar = df.loc[m_winter_non_solar, "Power Injected to SECI @ POC"].sum() / 1000.0
            
            seasonal_kpi_data = {
                "Category": ["Peak Time Evening", "Summer Solar Hours", "Summer Non-Solar Hours", "Winter Solar Hours", "Winter Non-Solar Hours"],
                "Start Hour": ["19.00", "06.00", "18.00", "07.00", "17.00"], "End Hour": ["01.00", "18.00", "06.00", "17.00", "07.00"],
                "Demand (MUs)": [demand_peak_eve, demand_summer_solar, demand_summer_non_solar, demand_winter_solar, demand_winter_non_solar],
                "Supply (MUs)": [supply_peak_eve, supply_summer_solar, supply_summer_non_solar, supply_winter_solar, supply_winter_non_solar],
                "Compliance (%)": [
                    (supply_peak_eve / demand_peak_eve * 100.0) if demand_peak_eve > 0 else 0.0,
                    (supply_summer_solar / demand_summer_solar * 100.0) if demand_summer_solar > 0 else 0.0,
                    (supply_summer_non_solar / demand_summer_non_solar * 100.0) if demand_summer_non_solar > 0 else 0.0,
                    (supply_winter_solar / demand_winter_solar * 100.0) if demand_winter_solar > 0 else 0.0,
                    (supply_winter_non_solar / demand_winter_non_solar * 100.0) if demand_winter_non_solar > 0 else 0.0,
                ],
                "Min": [90.00, 50.00, 70.00, 50.00, 70.00], "Max": [100.00, 60.00, 100.00, 60.00, 100.00]
            }
            df_seasonal_compliance = pd.DataFrame(seasonal_kpi_data)

            # Live Preview
            df_display = df.copy()
            numeric_cols = df_display.select_dtypes(include=[np.number]).columns.drop(["Availability", "DFR"])
            df_display[numeric_cols] = df_display[numeric_cols].round(2)
            df_display["Availability"] = df_display["Availability"].round(4)
            df_display["DFR"] = (df_display["DFR"] * 100).astype(str) + "%"
            
            st.markdown("### 📋 Live Data Preview Table (Active Mapping Confirmed)")
            clean_display_columns = ["Date", "Hour", "Time", "Peak/Non Peak", "DFR", "Availability", "DFR2", "Peak Demand @ POC (MW)", "Actual Peak Demand @ POC (MW)", "WIND Power @ 33 kV", "Solar Power @ 33 kV", "SOC", "Power Injected to SECI @ POC"]
            st.dataframe(df_display[clean_display_columns], use_container_width=True, height=400)
            
            # Master Compilation
            df_daily = df.groupby("Date").agg({"WIND Power @ 33 kV": "sum", "Solar Power @ 33 kV": "sum", "Total Power @ POC": "sum", "Charging Energy": "sum", "Discharging Energy": "sum", "Power Injected to SECI @ POC": "sum", "Energy to Exchange @ POC": "sum", "Unsold Energy @ POC": "sum", "BESS Standby Aux": "sum"}).reset_index()
            df_daily["Daily Average Availability"] = df_daily["Power Injected to SECI @ POC"] / (connectivity_limit_mwh * 96)
            
            datetime_series = pd.to_datetime(df["Date"], format="%d-%m-%Y")
            df["Month_Sort_Key"] = datetime_series.dt.to_period("M")
            df["Month-Year"] = datetime_series.dt.strftime("%B %Y")
            df_monthly = df.groupby(["Month_Sort_Key", "Month-Year"]).agg({"WIND Power @ 33 kV": "sum", "Solar Power @ 33 kV": "sum", "Total Power @ POC": "sum", "Charging Energy": "sum", "Discharging Energy": "sum", "Power Injected to SECI @ POC": "sum", "Energy to Exchange @ POC": "sum", "Unsold Energy @ POC": "sum", "BESS Standby Aux": "sum"}).reset_index().sort_values("Month_Sort_Key").drop(columns=["Month_Sort_Key"])
            
            total_hours = len(df) * 0.25
            sol_gen_mu = df["Solar Power @ 33 kV"].sum() / 1000.0
            wind_gen_mu = df["WIND Power @ 33 kV"].sum() / 1000.0
            to_bess_mu = df["Charging Energy"].sum() / ac_charging_eff / 1000.0
            from_bess_mu = abs(df["Charging Energy"].sum()) * ac_discharging_eff_pcs / 1000.0
            gross_33kv_mu = (sol_gen_mu + wind_gen_mu) - to_bess_mu + from_bess_mu
            net_poc_mu = df["Power Available for Grid at 220 KV @ POC"].sum() / 1000.0
            demand_poc_mu = df["Peak Demand @ POC (MW)"].sum() / 1000.0
            seci_poc_mu = df["Power Injected to SECI @ POC"].sum() / 1000.0
            peak_shortfall_mu = df.loc[df["Peak/Non Peak"] == 1, "DFR2"].sum() / 1000.0
            non_peak_shortfall_mu = df.loc[df["Peak/Non Peak"] == 0, "DFR2"].sum() / 1000.0
            
            df_main_summary = pd.DataFrame({
                "Description": ["Solar Capacity", "Wind Capacity", "PCSS Rating", "BESS Rating", "SECI Connectivity", "Exchange Connectivity", "Gross Solar Generation @ 33 kv PS", "Gross Wind Generation @ 33 kv PS", "Gross Hybrid Generation @ 33 kv P", "Energy to BESS from Hybrid @ 33 K", "Energy from BESS @ 33 KV", "Gross Generation @ 33 KV", "Net Generation @ POC", "Peak Demand @ POC", "Energy to SECI @ POC", "Energy to Exchange @ POC", "Unsold Energy @ POC", "Peak Shorfall", "Non Peak Shorfall", "Monthly Shortfall", "Annual Shortfall"],
                "Unit": ["MW", "MW", "MW", "MWh", "MW", "MW", "MU", "MU", "MU", "MU", "MU", "MU", "MU", "MU", "MU", "MU", "MU", "MU", "MU", "MU", "MU"],
                "Year 1": [solar_capacity, wind_capacity, pcs_capacity, bess_installed_capacity, grid_connectivity, exchange_connectivity, sol_gen_mu, wind_gen_mu, sol_gen_mu+wind_gen_mu, to_bess_mu, from_bess_mu, gross_33kv_mu, net_poc_mu, demand_poc_mu, seci_poc_mu, df["Energy to Exchange @ POC"].sum()/1000.0, df["Unsold Energy @ POC"].sum()/1000.0, peak_shortfall_mu, non_peak_shortfall_mu, 0.0, max(0.0, demand_poc_mu - seci_poc_mu)]
            })
            
            gross_solar_cuf = (sol_gen_mu * 1000.0) / (solar_capacity * total_hours) * 100.0 if solar_capacity > 0 else 0.0
            gross_wind_cuf = (wind_gen_mu * 1000.0) / (wind_capacity * total_hours) * 100.0 if wind_capacity > 0 else 0.0
            df_factor_summary = pd.DataFrame({
                "Description": ["Solar CuF", "Wind CuF", "RTC Compliance", "Exchange Exposure"],
                "Unit": ["%", "%", "%", "%"],
                "Gross": [f"{gross_solar_cuf:.2f}%", f"{gross_wind_cuf:.2f}%", f"{((seci_poc_mu / grid_connectivity) / 8.766 * 100.0):.2f}%", f"{((df['Energy to Exchange @ POC'].sum() / 1000.0) / max(1e-6, net_poc_mu)) * 100.0:.2f}%"],
                "Net": [f"{gross_solar_cuf * solar_hv_loss_factor:.2f}%", f"{gross_wind_cuf * wind_hv_loss_factor:.2f}%", "", ""]
            })
            
            st.markdown("### 📊 Performance Results Overview")
            c1, c2 = st.columns([2, 1.5])
            with c1: st.dataframe(df_main_summary.round(2), use_container_width=True, height=550)
            with c2: st.dataframe(df_factor_summary, use_container_width=True)
            st.markdown("### 🗓️ Seasonal & Peak Compliance Breakup")
            st.dataframe(df_seasonal_compliance, use_container_width=True)
                
            excel_buffer = io.BytesIO()
            with pd.ExcelWriter(excel_buffer, engine="openpyxl") as writer:
                df.round(3).to_excel(writer, sheet_name="15-Min Energy Analytics", index=False)
                df_daily.round(3).to_excel(writer, sheet_name="Daily Summary Matrix", index=False)
                df_monthly.round(3).to_excel(writer, sheet_name="Monthly Operational Trends", index=False)
                df_main_summary.round(2).to_excel(writer, sheet_name="Executive KPI Dashboard", startrow=1, index=False)
                df_factor_summary.to_excel(writer, sheet_name="Executive KPI Dashboard", startrow=25, index=False)
                df_seasonal_compliance.round(2).to_excel(writer, sheet_name="Executive KPI Dashboard", startrow=33, index=False)
                for sheet_key in writer.sheets:
                    ws = writer.sheets[sheet_key]
                    ws.views.sheetView[0].showGridLines = True
                    for col in ws.columns: ws.column_dimensions[get_column_letter(col[0].column)].width = max(max(len(str(cell.value or '')) for cell in col) + 3, 13)
            
            st.download_button("💾 Download Full Image-Matched Report (V3)", data=excel_buffer.getvalue(), file_name="PSA_RTC_Power_Model_V3.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            
        except Exception as e:
            st.error(f"Error processing data mapping pipelines: {e}")
    else:
        st.info("👆 Please upload the raw data Excel file in the sidebar to proceed.")


# =========================================================================================
# SUBLAYER 2: PPA RTC Power Model V2 (Standard DFR)
# =========================================================================================
def run_model_2_rtc_v2():
    st.title("⚡ PPA RTC Power Model (Standard DFR Engine)")

    st.sidebar.header("📁 Raw Data Upload")
    # CHANGED: Replaced local text path with a file uploader
    uploaded_file = st.sidebar.file_uploader("Upload Raw Data Excel File (.xlsx, .xlsm)", type=["xlsx", "xlsm"], key="m2_file")
    sheet_name = st.sidebar.text_input("Sheet Name", value="Raw Data", key="m2_sheet")

    wind_col_letter = st.sidebar.text_input("📊 WIND Per-Unit (PU) Column Letter", value="C", key="m2_wcol")
    solar_col_letter = st.sidebar.text_input("☀️ SOLAR Per-Unit (PU) Column Letter", value="D", key="m2_scol")
    data_start_row = st.sidebar.number_input("🔢 Starting Row for Raw Data (1-based)", value=6, min_value=1, step=1, key="m2_row")

    st.sidebar.markdown("---")
    st.sidebar.header("⏰ Peak Hours Configuration")
    enable_peak_hours = st.sidebar.checkbox("Enable Peak/Non-Peak Brackets", value=True, key="m2_peak_cb")

    if enable_peak_hours:
        morn_peak_start, morn_peak_end = st.sidebar.slider("Morning Peak Window", 0.0, 24.0, (0.0, 0.0), step=0.25, format="%.2f Hours", key="m2_mpeak")
        eve_peak_start, eve_peak_end = st.sidebar.slider("Evening Peak Window", 0.0, 24.0, (19.0, 21.0), step=0.25, format="%.2f Hours", key="m2_epeak")
    else:
        st.sidebar.info("ℹ️ Peak hours disabled. Poora operational spectrum All Non-Peak (0) par locked hai.")

    st.sidebar.markdown("---")
    st.sidebar.header("📊 Scenario & Probability Controls")
    wind_probability_scenario = st.sidebar.selectbox("Wind Probability Scenario (C34)", options=["P50", "P75", "P90"], index=0, key="m2_wps")
    solar_probability_scenario = st.sidebar.selectbox("Solar Probability Scenario (C38)", options=["P50", "P75", "P90"], index=0, key="m2_sps")
    p90_factor = st.sidebar.number_input("P90 Factor Value (C33)", value=0.924, step=0.001, format="%.3f", key="m2_p90")
    all_loss_factor = st.sidebar.number_input("Wind All Loss Factor (C43)", value=0.0050, step=0.05, format="%.2f", key="m2_loss")

    st.sidebar.markdown("---")
    st.sidebar.header("⚙️ Dynamic Plant Specifications")
    grid_connectivity = st.sidebar.number_input("Grid Connectivity Limit (MW - C1)", value=50.0, step=5.0, format="%.1f", key="m2_gc")
    exchange_connectivity = st.sidebar.number_input("Exchange Connectivity Limit (MW - C2)", value=25.0, step=5.0, format="%.1f", key="m2_ec")
    solar_capacity = st.sidebar.number_input("Solar Capacity (MW)", value=79.2, step=1.0, format="%.1f", key="m2_scap")
    wind_capacity = st.sidebar.number_input("Wind Capacity (MW)", value=49.5, step=1.0, format="%.1f", key="m2_wcap")
    pcs_capacity = st.sidebar.number_input("PCS Total Capacity (MW - C11)", value=85.0, step=1.0, format="%.1f", key="m2_pcs")
    bess_installed_capacity = st.sidebar.number_input("BESS Installed Capacity (MWh - C12)", value=170.0, step=1.0, format="%.1f", key="m2_bic")
    bess_usable_capacity = st.sidebar.number_input("BESS Usable Capacity (MWh - C14)", value=163.0, step=1.0, format="%.1f", key="m2_buc")
    baseline_peak_demand = st.sidebar.number_input("Baseline Demand Value @ POC (MW)", value=12.50, step=0.5, format="%.2f", key="m2_bpd")

    st.sidebar.markdown("---")
    st.sidebar.header("📉 Variable Loss & Efficiency Factors")
    solar_hv_loss_factor = st.sidebar.number_input("Solar HV Loss Factor (D22)", value=0.9626, step=0.001, format="%.4f", key="m2_shv")
    wind_hv_loss_factor = st.sidebar.number_input("Wind HV Loss Factor (E22)", value=0.9723, step=0.001, format="%.4f", key="m2_whv")
    ac_charging_eff = st.sidebar.number_input("AC Charging Efficiency (C22)", value=0.936, step=0.001, format="%.3f", key="m2_acc")
    total_charging_eff = st.sidebar.number_input("Total Charging Efficiency (C24)", value=0.902, step=0.001, format="%.3f", key="m2_tce")
    ac_discharging_eff_pcs = st.sidebar.number_input("AC Discharging Eff at PCS Terminal (C23)", value=0.936, step=0.001, format="%.3f", key="m2_acd")
    total_discharging_eff = st.sidebar.number_input("Total Discharging Efficiency (C25)", value=0.908, step=0.001, format="%.3f", key="m2_tde")
    ac_discharging_eff_poc = st.sidebar.number_input("AC Discharging Eff till POC (C27)", value=0.874, step=0.001, format="%.3f", key="m2_acdpoc")
    solar_degradation_factor = st.sidebar.number_input("Solar Degradation Factor (C41)", value=0.856, step=0.005, format="%.3f", key="m2_sdeg")
    wind_degradation_factor = st.sidebar.number_input("Wind Degradation Factor (C40)", value=0.990, step=0.005, format="%.3f", key="m2_wdeg")

    wind_idx = column_index_from_string(wind_col_letter) - 1
    solar_idx = column_index_from_string(solar_col_letter) - 1

    st.markdown("### 🗺️ Raw Input Sheet Mapping Guidelines")
    mapping_data = {
        "Excel Expected Profile Component": ["WIND Generation Per-Unit (PU) Data stream", "SOLAR Generation Per-Unit (PU) Data stream", "Data Starting Coordinate Point"],
        "Mapped Target Column Letter": [wind_col_letter.upper(), solar_col_letter.upper(), f"Row {data_start_row} (1-based)"],
        "Zero-Based Python Index Mapping": [f"Index {wind_idx}", f"Index {solar_idx}", f"Skipping first {data_start_row - 1} row(s)"]
    }
    st.table(pd.DataFrame(mapping_data))

    # CHANGED: Check if file is uploaded instead of local path existence
    if uploaded_file is not None:
        try:
            df_raw = pd.read_excel(uploaded_file, sheet_name=sheet_name, skiprows=int(data_start_row - 1), header=None, engine="openpyxl")
            df = pd.DataFrame()
            
            generated_timestamps = pd.date_range(start="1990-01-01 00:00", periods=len(df_raw), freq="15min")
            df["Date"] = generated_timestamps.strftime("%d-%m-%Y")
            df["Hour"] = generated_timestamps.hour
            df["Time"] = generated_timestamps.strftime("%H:%M")
            
            if enable_peak_hours:
                decimal_hours = generated_timestamps.hour + generated_timestamps.minute / 60.0
                is_morn_peak = (morn_peak_start != morn_peak_end) & (decimal_hours >= morn_peak_start) & (decimal_hours <= morn_peak_end)
                is_eve_peak = (eve_peak_start != eve_peak_end) & (decimal_hours >= eve_peak_start) & (decimal_hours <= eve_peak_end)
                df["Peak/Non Peak"] = np.where(is_morn_peak | is_eve_peak, 1, 0)
            else:
                df["Peak/Non Peak"] = 0
            
            df["DFR"] = np.where(df["Peak/Non Peak"] == 0, 0.70, 0.90)
            df["Peak Demand @ POC (MW)"] = baseline_peak_demand
            
            if wind_idx >= len(df_raw.columns) or solar_idx >= len(df_raw.columns):
                st.error(f"🚨 Configuration Error: Mapped Column Letters fall out of bounds!")
                st.stop()
                
            raw_wind_pu = pd.to_numeric(df_raw.iloc[:, wind_idx], errors='coerce').fillna(0.0)
            raw_solar_pu = pd.to_numeric(df_raw.iloc[:, solar_idx], errors='coerce').fillna(0.0)
            
            wind_multiplier = 1.0 if wind_probability_scenario in ["P50", "P75"] else p90_factor
            converted_wind_pu = (raw_wind_pu * wind_multiplier) * (1.0 + all_loss_factor)
            
            solar_multiplier = 1.0 if solar_probability_scenario in ["P50", "P75"] else p90_factor
            converted_solar_pu = raw_solar_pu * solar_multiplier
            
            df["WIND Power @ 33 kV"] = converted_wind_pu * (wind_capacity / 4.0) * wind_degradation_factor
            df["Solar Power @ 33 kV"] = converted_solar_pu * (solar_capacity / 4.0) * solar_degradation_factor
            df["Total Power @ POC"] = df["WIND Power @ 33 kV"] + df["Solar Power @ 33 kV"]
            
            df["Power Beyond PPA Demand"] = np.where(df["Total Power @ POC"] > df["Peak Demand @ POC (MW)"], df["Total Power @ POC"] - df["Peak Demand @ POC (MW)"], 0.0)
            df["Power Available for Charging POC"] = (df["Total Power @ POC"] * solar_hv_loss_factor) - df["Peak Demand @ POC (MW)"]
            df["Is it Charge (C)/ Discharge (D)"] = np.where(df["Power Available for Charging POC"] < 0.0, "D", "C")
            df["If D, then Energy Discharges from PCS"] = np.where(df["Is it Charge (C)/ Discharge (D)"] == "D", df["Power Available for Charging POC"] / ac_discharging_eff_poc, 0.0)
            
            pcs_dis_limit_mwh = (pcs_capacity / ac_discharging_eff_pcs) / 4.0
            df["Discharging Power Limited to PCS Rating"] = np.where(df["If D, then Energy Discharges from PCS"].abs() > pcs_dis_limit_mwh, -pcs_dis_limit_mwh, df["If D, then Energy Discharges from PCS"])
            
            size = len(df)
            energy_goes_for_ess, charging_power_limited_to_ess, daily_charging_limit, soc_tracker = np.zeros(size), np.zeros(size), np.zeros(size), np.zeros(size)
            charge_discharge_increment, charging_energy, discharging_energy, abs_of_increment = np.zeros(size), np.zeros(size), np.zeros(size), np.zeros(size)
            power_beyond_full_charging, bess_standby_aux = np.zeros(size), np.zeros(size)
            
            current_soc = bess_usable_capacity
            dates, modes = df["Date"].values, df["Is it Charge (C)/ Discharge (D)"].values
            total_power_poc, power_beyond_ppa = df["Total Power @ POC"].values, df["Power Beyond PPA Demand"].values
            discharging_power_limited = df["Discharging Power Limited to PCS Rating"].values
            
            pcs_chg_limit_mwh = (pcs_capacity * ac_charging_eff) / 4.0
            max_daily_charging = bess_usable_capacity * 2.0
            
            for i in range(size):
                prev_daily_limit = 0.0 if i == 0 or dates[i] != dates[i-1] else daily_charging_limit[i-1]
                prev_soc = current_soc
                
                energy_goes_for_ess[i] = power_beyond_ppa[i] * total_charging_eff if (prev_soc < bess_usable_capacity) and (modes[i] == "C") else 0.0
                charging_power_limited_to_ess[i] = pcs_chg_limit_mwh if energy_goes_for_ess[i] > pcs_chg_limit_mwh else energy_goes_for_ess[i]
                daily_charging_limit[i] = min(prev_daily_limit + charging_power_limited_to_ess[i], max_daily_charging) if modes[i] == "C" else prev_daily_limit
                
                trial_soc = prev_soc + discharging_power_limited[i] + charging_power_limited_to_ess[i]
                if (modes[i] == "D") and (trial_soc < 0.0): current_soc = 0.0
                elif trial_soc > bess_usable_capacity: current_soc = bess_usable_capacity
                else: current_soc = prev_soc + discharging_power_limited[i] if (daily_charging_limit[i] >= max_daily_charging) else trial_soc
                        
                soc_tracker[i] = current_soc
                charge_discharge_increment[i] = current_soc - prev_soc
                charging_energy[i] = charge_discharge_increment[i] if charge_discharge_increment[i] > 0.0 else 0.0
                discharging_energy[i] = charge_discharge_increment[i] if charge_discharge_increment[i] < 0.0 else 0.0
                abs_of_increment[i] = abs(charge_discharge_increment[i])
                surplus_ac = total_power_poc[i] - (abs_of_increment[i] / ac_charging_eff)
                power_beyond_full_charging[i] = surplus_ac if surplus_ac > 0.0 else 0.0
                bess_standby_aux[i] = (3.5 * bess_installed_capacity / 4.0) / 1000.0 if charge_discharge_increment[i] == 0.0 else 0.0
                
            df["If C, Energy Goes for ESS"] = energy_goes_for_ess
            df["Charging Power Limited to ESS Capacity"] = charging_power_limited_to_ess
            df["Daily Charging Limit"] = daily_charging_limit
            df["SOC"] = soc_tracker
            df["Charge/Discharge Increment"] = charge_discharge_increment
            df["Charging Energy"] = charging_energy
            df["Discharging Energy"] = discharging_energy
            df["ABS of 12"] = abs_of_increment
            df["Power Beyond Full Charging"] = power_beyond_full_charging
            df["BESS Standby Aux"] = bess_standby_aux
            
            df["Power Available for Grid at 220 KV @ POC"] = np.where(
                df["Is it Charge (C)/ Discharge (D)"] == "D",
                df["Total Power @ POC"] - (df["Charge/Discharge Increment"] * total_discharging_eff),
                df["Power Beyond Full Charging"]
            ) * solar_hv_loss_factor - df["BESS Standby Aux"]
            
            connectivity_limit_mwh = grid_connectivity / 4.0
            df["Power Injected to SECI @ POC"] = np.where(df["Power Available for Grid at 220 KV @ POC"] > connectivity_limit_mwh, connectivity_limit_mwh, df["Power Available for Grid at 220 KV @ POC"])
            df["Excess Power beyond evacuation level @ POC"] = df["Power Available for Grid at 220 KV @ POC"] - df["Power Injected to SECI @ POC"]
            connectivity_exchange_limit_mwh = exchange_connectivity / 4.0
            df["Energy to Exchange @ POC"] = np.where(df["Excess Power beyond evacuation level @ POC"] < connectivity_exchange_limit_mwh, df["Excess Power beyond evacuation level @ POC"], connectivity_exchange_limit_mwh)
            df["Unsold Energy @ POC"] = df["Excess Power beyond evacuation level @ POC"] - df["Energy to Exchange @ POC"]
            df["Availability"] = df["Power Injected to SECI @ POC"] / connectivity_limit_mwh
            df["DFR2"] = np.where(df["Availability"] < df["DFR"], (connectivity_limit_mwh * df["DFR"]) - df["Power Available for Grid at 220 KV @ POC"], 0.0)
            
            # Rendering Display
            df_display = df.copy()
            numeric_cols = df_display.select_dtypes(include=[np.number]).columns.drop(["Availability", "DFR"])
            df_display[numeric_cols] = df_display[numeric_cols].round(2)
            df_display["Availability"] = df_display["Availability"].round(4)
            df_display["DFR"] = (df_display["DFR"] * 100).astype(str) + "%"
            
            st.markdown("### 📋 Live Data Preview Table (Active Mapping Confirmed)")
            clean_display_columns = ["Date", "Hour", "Time", "Peak/Non Peak", "DFR", "Availability", "DFR2", "WIND Power @ 33 kV", "Solar Power @ 33 kV", "Total Power @ POC", "SOC", "Power Injected to SECI @ POC", "Energy to Exchange @ POC"]
            st.dataframe(df_display[clean_display_columns], use_container_width=True, height=400)
            
            df_daily = df.groupby("Date").agg({"WIND Power @ 33 kV": "sum", "Solar Power @ 33 kV": "sum", "Total Power @ POC": "sum", "Charging Energy": "sum", "Discharging Energy": "sum", "Power Injected to SECI @ POC": "sum", "Energy to Exchange @ POC": "sum", "Unsold Energy @ POC": "sum", "BESS Standby Aux": "sum"}).reset_index()
            df_daily["Daily Average Availability"] = df_daily["Power Injected to SECI @ POC"] / (connectivity_limit_mwh * 96)
            
            datetime_series = pd.to_datetime(df["Date"], format="%d-%m-%Y")
            df["Month_Sort_Key"] = datetime_series.dt.to_period("M")
            df["Month-Year"] = datetime_series.dt.strftime("%B %Y")
            df_monthly = df.groupby(["Month_Sort_Key", "Month-Year"]).agg({"WIND Power @ 33 kV": "sum", "Solar Power @ 33 kV": "sum", "Total Power @ POC": "sum", "Charging Energy": "sum", "Discharging Energy": "sum", "Power Injected to SECI @ POC": "sum", "Energy to Exchange @ POC": "sum", "Unsold Energy @ POC": "sum", "BESS Standby Aux": "sum"}).reset_index().sort_values("Month_Sort_Key").drop(columns=["Month_Sort_Key"])
            
            total_hours = len(df) * 0.25
            sol_gen_mu = df["Solar Power @ 33 kV"].sum() / 1000.0
            wind_gen_mu = df["WIND Power @ 33 kV"].sum() / 1000.0
            to_bess_mu = df["Charging Energy"].sum() / ac_charging_eff / 1000.0
            from_bess_mu = abs(df["Charging Energy"].sum()) * ac_discharging_eff_pcs / 1000.0
            gross_33kv_mu = (sol_gen_mu + wind_gen_mu) - to_bess_mu + from_bess_mu
            net_poc_mu = df["Power Available for Grid at 220 KV @ POC"].sum() / 1000.0
            demand_poc_mu = (df["Peak Demand @ POC (MW)"].sum()) / 1000.0
            seci_poc_mu = df["Power Injected to SECI @ POC"].sum() / 1000.0
            peak_shortfall_mu = df.loc[df["Peak/Non Peak"] == 1, "DFR2"].sum() / 1000.0
            non_peak_shortfall_mu = df.loc[df["Peak/Non Peak"] == 0, "DFR2"].sum() / 1000.0
            
            df_main_summary = pd.DataFrame({
                "Description": ["Solar Capacity", "Wind Capacity", "PCSS Rating", "BESS Rating", "SECI Connectivity", "Exchange Connectivity", "Gross Solar Generation @ 33 kv PS", "Gross Wind Generation @ 33 kv PS", "Gross Hybrid Generation @ 33 kv P", "Energy to BESS from Hybrid @ 33 K", "Energy from BESS @ 33 KV", "Gross Generation @ 33 KV", "Net Generation @ POC", "Peak Demand @ POC", "Energy to SECI @ POC", "Energy to Exchange @ POC", "Unsold Energy @ POC", "Peak Shorfall", "Non Peak Shorfall", "Monthly Shortfall", "Annual Shortfall"],
                "Unit": ["MW", "MW", "MW", "MWh", "MW", "MW", "MU", "MU", "MU", "MU", "MU", "MU", "MU", "MU", "MU", "MU", "MU", "MU", "MU", "MU", "MU"],
                "Year 1": [solar_capacity, wind_capacity, pcs_capacity, bess_installed_capacity, grid_connectivity, exchange_connectivity, sol_gen_mu, wind_gen_mu, sol_gen_mu+wind_gen_mu, to_bess_mu, from_bess_mu, gross_33kv_mu, net_poc_mu, demand_poc_mu, seci_poc_mu, df["Energy to Exchange @ POC"].sum()/1000.0, df["Unsold Energy @ POC"].sum()/1000.0, peak_shortfall_mu, non_peak_shortfall_mu, 0.0, max(0.0, demand_poc_mu - seci_poc_mu)]
            })
            
            gross_solar_cuf = (sol_gen_mu * 1000.0) / (solar_capacity * total_hours) * 100.0 if solar_capacity > 0 else 0.0
            gross_wind_cuf = (wind_gen_mu * 1000.0) / (wind_capacity * total_hours) * 100.0 if wind_capacity > 0 else 0.0
            df_factor_summary = pd.DataFrame({
                "Description": ["Solar CuF", "Wind CuF", "RTC Compliance", "Exchange Exposure"],
                "Unit": ["%", "%", "%", "%"],
                "Gross": [f"{gross_solar_cuf:.2f}%", f"{gross_wind_cuf:.2f}%", f"{((seci_poc_mu / grid_connectivity) / 8.766 * 100.0):.2f}%", f"{((df['Energy to Exchange @ POC'].sum() / 1000.0) / max(1e-6, net_poc_mu)) * 100.0:.2f}%"],
                "Net": [f"{gross_solar_cuf * solar_hv_loss_factor:.2f}%", f"{gross_wind_cuf * wind_hv_loss_factor:.2f}%", "", ""]
            })
            
            st.markdown("### 📊 Performance Results Overview")
            c1, c2 = st.columns([2, 1.5])
            with c1: st.dataframe(df_main_summary.round(2), use_container_width=True, height=550)
            with c2: st.dataframe(df_factor_summary, use_container_width=True)
                
            excel_buffer = io.BytesIO()
            with pd.ExcelWriter(excel_buffer, engine="openpyxl") as writer:
                df.round(3).to_excel(writer, sheet_name="15-Min Energy Analytics", index=False)
                df_daily.round(3).to_excel(writer, sheet_name="Daily Summary Matrix", index=False)
                df_monthly.round(3).to_excel(writer, sheet_name="Monthly Operational Trends", index=False)
                df_main_summary.round(2).to_excel(writer, sheet_name="Executive KPI Dashboard", startrow=1, index=False)
                df_factor_summary.to_excel(writer, sheet_name="Executive KPI Dashboard", startrow=25, index=False)
                for sheet_key in writer.sheets:
                    ws = writer.sheets[sheet_key]
                    ws.views.sheetView[0].showGridLines = True
                    for col in ws.columns: ws.column_dimensions[get_column_letter(col[0].column)].width = max(max(len(str(cell.value or '')) for cell in col) + 3, 13)
            
            st.download_button("💾 Download Full Image-Matched Report (V2)", data=excel_buffer.getvalue(), file_name="PSA_RTC_Power_Model_V2.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            
        except Exception as e:
            st.error(f"Error processing data mapping pipelines: {e}")
    else:
        st.info("👆 Please upload the raw data Excel file in the sidebar to proceed.")


# =========================================================================================
# SUBLAYER 3: NHPC BESS Hybrid 25-Year Model
# =========================================================================================
def run_model_3_nhpc_25yr():
    st.title("⚡ NHPC BESS Hybrid 25-Year Model")

    with st.sidebar.expander("🔋 Core Plant & Grid Constants", expanded=True):
        solar_capacity = st.number_input("Solar Capacity (MW)", value=123.2, step=1.0, key="m3_scap")
        wind_capacity = st.number_input("Wind Capacity (MW)", value=145.2, step=1.0, key="m3_wcap")
        high_voltage_loss = st.number_input("Total High Voltage Loss (as decimal)", value=-0.016, step=0.005, format="%.3f", key="m3_hvl")
        connectivity = st.number_input("Grid Connectivity Limit (MW)", value=100.0, step=1.0, key="m3_conn")
        exchange_connectivity = st.number_input("Exchange Connectivity Limit (MW)", value=150.0, step=1.0, key="m3_exconn")
        peak_demand_poc = st.number_input("Peak Demand at POC (MW)", value=100.0, step=1.0, key="m3_pd")
        manual_AC_to_DC = st.number_input("AC_to_DC_eff(Charging)", value=0.94, step=0.01, key="m3_acdc")
        manual_DC_to_AC = st.number_input("DC_to_AC_eff(Discharging)", value=0.93, step=0.01, key="m3_dcac")
        bess_usable_capacity = st.number_input("BESS Usable Capacity (MWh)", value=119.8, step=1.0, key="m3_buc")
        pcs_rating = st.number_input("PCS Rating (MW)", value=65.0, step=1.0, key="m3_pcs")

    with st.sidebar.expander("📁 Efficiency & 25-Yr BESS Loader", expanded=True):
        # CHANGED: Replaced local text path with a file uploader
        uploaded_constants = st.file_uploader("Upload Constants/Loss Assumption Excel File", type=["xlsx", "xlsm"], key="m3_constfile")
        constants_sheet_name = st.text_input("Constants Sheet Name", value="Loss Assumption", key="m3_constsheet")
        
        AC_to_DC_eff = manual_AC_to_DC
        DC_to_AC_eff = manual_DC_to_AC
        bess_lifetime_capacities = [bess_usable_capacity] * 25
        
        if uploaded_constants is not None:
            df_const = None
            try:
                df_const = pd.read_excel(uploaded_constants, sheet_name=constants_sheet_name, header=None, engine="openpyxl")
            except Exception:
                try:
                    df_const = pd.read_excel(uploaded_constants, sheet_name=constants_sheet_name, header=None)
                except Exception as inner_e:
                    st.warning(f"Using manual configurations. Info: {inner_e}")
            
            if df_const is not None:
                try:
                    total_charge_eff = pd.to_numeric(df_const.iloc[6, 1], errors='coerce')     
                    total_discharge_eff = pd.to_numeric(df_const.iloc[7, 1], errors='coerce')  
                    
                    if total_charge_eff > 1.0: total_charge_eff /= 100.0
                    if total_discharge_eff > 1.0: total_discharge_eff /= 100.0
                    
                    AC_to_DC_eff = total_charge_eff
                    DC_to_AC_eff = total_discharge_eff
                    
                    parsed_bess_profile = []
                    for y in range(1, 26):
                        row_idx = 38 + (y - 1)  
                        val = pd.to_numeric(df_const.iloc[row_idx, 6], errors='coerce') 
                        if pd.isna(val): parsed_bess_profile.append(bess_usable_capacity)
                        elif val < 2.0: parsed_bess_profile.append(bess_usable_capacity * val)
                        else: parsed_bess_profile.append(val)
                    bess_lifetime_capacities = parsed_bess_profile
                    st.success("🎯 Efficiencies & 25-Year BESS Capacity Map Loaded!")
                except Exception as calc_e:
                    st.error(f"Structure index mapping mismatch: {calc_e}")
        else:
            st.warning("Constants file not uploaded. Running standard baseline profiles.")

    with st.sidebar.expander("📉 Resource Scale Modifiers", expanded=True):
        solar_p = st.selectbox("Solar Probability Factor Scenario", options=["P50", "P75", "P90"], index=0, key="m3_sp")
        wind_p = st.selectbox("Wind Probability Factor Scenario", options=["P50", "P75", "P90"], index=0, key="m3_wp")
        wind_p75_scale = st.number_input("Wind P75 Scaling Factor", value=0.965, step=0.005, format="%.3f", key="m3_wp75")
        wind_p90_scale = st.number_input("Wind P90 Scaling Factor (I3)", value=0.935, step=0.005, format="%.3f", key="m3_wp90")
        solar_p75_scale = st.number_input("Solar P75 Scaling Factor", value=0.965, step=0.005, format="%.3f", key="m3_sp75")
        solar_p90_scale = st.number_input("Solar P90 Scaling Factor (J3)", value=0.935, step=0.005, format="%.3f", key="m3_sp90")

        st.markdown("---")
        wind_deg = st.number_input("Wind Degradation Year Factor (D3)", value=1.000, step=0.01, format="%.3f", key="m3_wdyf")
        wind_reeval = st.number_input("Wind Re-evaluated Loss (I20) [as decimal]", value=-0.0187, step=0.001, format="%.4f", key="m3_wrl")
        wind_unavail = st.number_input("Wind 33KV Unavailability (I24) [as decimal]", value=0.0020, step=0.001, format="%.4f", key="m3_wua")

        st.markdown("---")
        solar_deg = st.number_input("Solar Degradation Year Factor (C3)", value=1.000, step=0.01, format="%.3f", key="m3_sdyf")
        solar_unavail = st.number_input("Solar Inverter Unavailability (H24) [as decimal]", value=0.0050, step=0.001, format="%.4f", key="m3_sua")
        solar_fixed_mult = st.number_input("Solar Fixed Multiplier Addition", value=0.0040, step=0.001, format="%.4f", key="m3_sfma")

    with st.sidebar.expander("📉 Lifetime Degradation Rates", expanded=True):
        solar_annual_deg = st.number_input("Solar Annual Degradation Rate", value=0.0045, format="%.4f", key="m3_sadr")
        wind_annual_deg = st.number_input("Wind Annual Degradation Rate", value=0.0002, format="%.4f", key="m3_wadr")

    with st.sidebar.expander("🔌 Plant Configuration Mode", expanded=False):
        plant_config = st.selectbox("Select Active Generation Sources", options=["Solar + Wind + BESS", "Solar + BESS", "Wind + BESS"], index=0, key="m3_pconf")

    with st.sidebar.expander("⏰ Dynamic Peak Window Settings", expanded=False):
        morning_peak_range = st.slider("Morning Peak Bracket (Hours)", min_value=0.0, max_value=12.0, value=(7.0, 9.0), step=0.25, key="m3_mpr")
        evening_peak_range = st.slider("Evening Peak Bracket (Hours)", min_value=12.0, max_value=24.0, value=(19.0, 21.0), step=0.25, key="m3_epr")

    with st.sidebar.expander("📁 Data Source & Column Setup", expanded=True):
        # CHANGED: Replaced local text path with a file uploader
        uploaded_file = st.file_uploader("Upload Main Raw Data Excel File", type=["xlsx", "xlsm"], key="m3_mainfile")
        sheet_name = st.text_input("Sheet Name", value="Sheet1", key="m3_sn")
        data_start_row = st.number_input("Data Starts on Row", value=2, min_value=1, step=1, key="m3_dsr")
        wind_col_letter = st.text_input("Wind Column Letter", value="C", key="m3_wcl")
        solar_col_letter = st.text_input("Solar Column Letter", value="D", key="m3_scl")

    wind_idx = column_index_from_string(wind_col_letter) - 1
    solar_idx = column_index_from_string(solar_col_letter) - 1

    if uploaded_file is not None:
        with st.spinner("Executing Independent P-Factor 25-Year Lifecycle Matrix..."):
            try:
                df_raw = load_data_nhpc(uploaded_file, sheet_name, data_start_row)
                df_base = pd.DataFrame()
                
                max_cols = len(df_raw.columns)
                if wind_idx < max_cols and solar_idx < max_cols:
                    generated_timestamps = pd.date_range(start="2026-01-01 00:00", periods=len(df_raw), freq="15min")
                    df_base["Time"] = generated_timestamps.strftime("%H:%M")
                    df_base["Month_Name"] = generated_timestamps.strftime("%B")
                    df_base["Month_Num"] = generated_timestamps.month
                    decimal_hours = generated_timestamps.hour + generated_timestamps.minute / 60.0
                    total_hours = len(df_base) / 4.0
                    
                    is_morning_peak = (decimal_hours >= morning_peak_range[0]) & (decimal_hours < morning_peak_range[1])
                    is_evening_peak = (decimal_hours >= evening_peak_range[0]) & (decimal_hours < evening_peak_range[1])
                    is_peak_time = is_morning_peak | is_evening_peak
                    
                    peak_demand_33kv = peak_demand_poc / (1.0 + high_voltage_loss)
                    df_base["Peak_Demand_POC_MW"] = np.where(is_peak_time, peak_demand_poc, 0.0)
                    df_base["Peak_Demand_33KV_MW"] = np.where(is_peak_time, peak_demand_33kv, 0.0)
                        
                    df_base["Raw_Solar"] = pd.to_numeric(df_raw.iloc[:, solar_idx], errors='coerce').fillna(0.0)
                    df_base["Raw_Wind"] = pd.to_numeric(df_raw.iloc[:, wind_idx], errors='coerce').fillna(0.0)

                    base_wind_mod = df_base["Raw_Wind"] * wind_deg * (1 + wind_reeval) * (1 - wind_unavail)
                    if wind_p == "P50": df_base["Raw_Wind_Modified"] = base_wind_mod
                    elif wind_p == "P75": df_base["Raw_Wind_Modified"] = base_wind_mod * wind_p75_scale
                    else: df_base["Raw_Wind_Modified"] = base_wind_mod * wind_p90_scale

                    base_solar_mod = df_base["Raw_Solar"] * solar_deg * (1 - solar_unavail)* (1 + solar_fixed_mult)
                    if solar_p == "P50": df_base["Raw_Solar_Modified"] = base_solar_mod 
                    elif solar_p == "P75": df_base["Raw_Solar_Modified"] = base_solar_mod * solar_p75_scale 
                    else: df_base["Raw_Solar_Modified"] = base_solar_mod * solar_p90_scale 

                    yearly_compiled_snapshots = {}
                    months_list = ["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]
                    df_year1_sample = None

                    for year in range(1, 26):
                        df_y = df_base.copy()
                        solar_deg_factor = 1.0 - (year - 1) * solar_annual_deg
                        wind_deg_factor = 1.0 - (year - 1) * wind_annual_deg
                        current_year_bess_cap = bess_lifetime_capacities[year - 1]
                        
                        scaled_solar_profile = df_y["Raw_Solar_Modified"] * solar_deg_factor
                        scaled_wind_profile = df_y["Raw_Wind_Modified"] * wind_deg_factor
                        
                        if plant_config == "Solar + Wind + BESS":
                            df_y["Solar_Generation_MW"] = scaled_solar_profile * solar_capacity
                            df_y["Wind_Generation_MW"] = scaled_wind_profile * wind_capacity
                        elif plant_config == "Solar + BESS":
                            df_y["Solar_Generation_MW"] = scaled_solar_profile * solar_capacity
                            df_y["Wind_Generation_MW"] = 0.0
                        elif plant_config == "Wind + BESS":
                            df_y["Solar_Generation_MW"] = 0.0
                            df_y["Wind_Generation_MW"] = scaled_wind_profile * wind_capacity

                        df_y["Total_RE_Generation_MW"] = df_y["Solar_Generation_MW"] + df_y["Wind_Generation_MW"]
                        limit_at_33kv = connectivity / (1 + high_voltage_loss)

                        df_y["Power_Beyond_SECI_Limit_MW"] = np.where(df_y["Total_RE_Generation_MW"] > limit_at_33kv, df_y["Total_RE_Generation_MW"] - limit_at_33kv, 0.0)
                        df_y["Energy_Available_for_BESS_MW"] = df_y["Total_RE_Generation_MW"] - df_y["Peak_Demand_33KV_MW"]
                        df_y["Is it Charge (C )/ Discharge (D)"] = np.where(df_y["Energy_Available_for_BESS_MW"] < 0, "D", "C")
                        df_y["If_D_then_BESS_Discharge_MW"] = np.where(df_y["Is it Charge (C )/ Discharge (D)"] == "D", df_y["Energy_Available_for_BESS_MW"] / DC_to_AC_eff, 0.0)

                        pcs_limit_dc = pcs_rating / DC_to_AC_eff
                        df_y["BESS_Discharge_MWh"] = np.where(df_y["If_D_then_BESS_Discharge_MW"].abs() > pcs_limit_dc, -pcs_limit_dc / 4.0, df_y["If_D_then_BESS_Discharge_MW"] / 4.0)

                        df_y["Block_ID"] = (df_y["Is it Charge (C )/ Discharge (D)"] != df_y["Is it Charge (C )/ Discharge (D)"].shift()).cumsum()
                        df_y["Excess_Energy_Upto_Next_Peak_MW"] = df_y[::-1].groupby("Block_ID")["Power_Beyond_SECI_Limit_MW"].cumsum()
                        df_y["Excess_Energy_Upto_Next_Peak_MW"] = np.where(df_y["Is it Charge (C )/ Discharge (D)"] == "D", 0.0, df_y["Excess_Energy_Upto_Next_Peak_MW"])

                        size = len(df_y)
                        if_c_charges, strategic_charges, energy_used_charging, c_limits, soc_tracker, soc_increments = np.zeros(size), np.zeros(size), np.zeros(size), np.zeros(size), np.zeros(size), np.zeros(size)
                        modes, p_beyonds, excess_peaks, total_res, d_limits = df_y["Is it Charge (C )/ Discharge (D)"].values, df_y["Power_Beyond_SECI_Limit_MW"].values, df_y["Excess_Energy_Upto_Next_Peak_MW"].values, df_y["Total_RE_Generation_MW"].values, df_y["BESS_Discharge_MWh"].values  
                        
                        current_soc = current_year_bess_cap 
                        
                        for i in range(size):
                            prev_soc = current_soc
                            k, p_beyond, excess_upto_peak, total_re, m = modes[i], p_beyonds[i], excess_peaks[i], total_res[i], d_limits[i]
                            
                            if_c_charge_mw = p_beyond * AC_to_DC_eff if (prev_soc < current_year_bess_cap) and (k == "C") else 0.0
                            if_c_charges[i] = if_c_charge_mw
                            
                            if (prev_soc < current_year_bess_cap) and (k == "C") and (excess_upto_peak > (current_year_bess_cap - prev_soc)): strategic_charge_mw = p_beyond * AC_to_DC_eff
                            elif prev_soc >= current_year_bess_cap: strategic_charge_mw = 0.0
                            else: strategic_charge_mw = total_re * AC_to_DC_eff
                            strategic_charges[i] = strategic_charge_mw
                            
                            energy_used_mw = strategic_charge_mw if k == "C" else 0.0
                            energy_used_charging[i] = energy_used_mw
                            r = (pcs_rating * AC_to_DC_eff) / 4.0 if energy_used_mw > (pcs_rating * AC_to_DC_eff) else energy_used_mw / 4.0
                            c_limits[i] = r
                            
                            if k == "D" and (prev_soc + m + r < 0): current_soc = 0.0
                            elif (prev_soc + m + r) > current_year_bess_cap: current_soc = current_year_bess_cap
                            else: current_soc = prev_soc + m + r 
                            soc_tracker[i], soc_increments[i] = current_soc, current_soc - prev_soc
                        
                        df_y["SOC"], df_y["Charge_Discharge_Increment_MWh"], df_y["ABS_Charge_Discharge_Increment_MWh"] = soc_tracker, soc_increments, np.abs(soc_increments)
                        df_y["BESS_Charge_Source"] = np.where(energy_used_charging > 0, np.where(p_beyonds > 0, "E", "M"), "0")

                        df_y["Power_to_BESS_33KV_MWh"] = np.where(df_y["Charge_Discharge_Increment_MWh"] > 0, df_y["Charge_Discharge_Increment_MWh"] / AC_to_DC_eff, 0.0)
                        df_y["Power_from_BESS_33KV_MWh"] = np.where(df_y["Charge_Discharge_Increment_MWh"] < 0, df_y["Charge_Discharge_Increment_MWh"].abs() * DC_to_AC_eff, 0.0)

                        remaining_generation = df_y["Total_RE_Generation_MW"] - (df_y["ABS_Charge_Discharge_Increment_MWh"] * 4.0 / AC_to_DC_eff)
                        df_y["Unutilised_Beyond_Full_Charging_MW"] = np.where(df_y["BESS_Charge_Source"] == "0", df_y["Total_RE_Generation_MW"], np.where(remaining_generation > 0, remaining_generation, 0.0))

                        df_y["Energy_Available_for_Grid_POC_MW"] = np.where(df_y["Is it Charge (C )/ Discharge (D)"] == "D", df_y["Total_RE_Generation_MW"] - (df_y["Charge_Discharge_Increment_MWh"] * 4.0 * DC_to_AC_eff), df_y["Unutilised_Beyond_Full_Charging_MW"]) * (1 + high_voltage_loss)
                        df_y["Energy_Injected_SECI_POC_MW"] = np.where(df_y["Energy_Available_for_Grid_POC_MW"] > connectivity, connectivity, df_y["Energy_Available_for_Grid_POC_MW"])
                        df_y["Excess_Energy_Beyond_Evacuation_POC_MW"] = (df_y["Energy_Available_for_Grid_POC_MW"] - df_y["Energy_Injected_SECI_POC_MW"])
                        df_y["Energy_to_Exchange_POC_MW"] = np.where(df_y["Excess_Energy_Beyond_Evacuation_POC_MW"] < exchange_connectivity, df_y["Excess_Energy_Beyond_Evacuation_POC_MW"], exchange_connectivity)
                        df_y["Unsold_Energy_POC_MW"] = (df_y["Excess_Energy_Beyond_Evacuation_POC_MW"] - df_y["Energy_to_Exchange_POC_MW"])
                        
                        df_y["Net_Demand_MW"] = np.where(df_y["Peak_Demand_POC_MW"] > connectivity, connectivity, 0.0)
                        df_y["Energy_to_Grid_During_Peak_MW"] = np.where(df_y["Peak_Demand_POC_MW"] > 0, df_y["Energy_Injected_SECI_POC_MW"], 0.0)

                        monthly_shortfall = {}
                        for month in range(1, 13):
                            month_df = df_y[df_y["Month_Num"] == month]
                            net_demand_sum = month_df["Net_Demand_MW"].sum()
                            energy_grid_sum = month_df["Energy_to_Grid_During_Peak_MW"].sum()
                            required_energy = (net_demand_sum / 4.0) * 0.90
                            actual_energy = energy_grid_sum / 4.0
                            monthly_shortfall[month] = max(0.0, required_energy - actual_energy)

                        df_y["Shortfall_MWh"] = np.where(is_peak_time & (df_y["Net_Demand_MW"] > 0), df_y["Month_Num"].map(monthly_shortfall) / (df_y.groupby("Month_Num")["Net_Demand_MW"].transform("count")), 0.0)
                        sf = round(sum(monthly_shortfall.values()) / 1000.0, 2)

                        g_solar, g_wind = df_y["Solar_Generation_MW"].sum() / 4000.0, df_y["Wind_Generation_MW"].sum() / 4000.0
                        to_b, from_b = df_y["Power_to_BESS_33KV_MWh"].sum() / 1000.0, df_y["Power_from_BESS_33KV_MWh"].sum() / 1000.0
                        n_poc, seci_poc = df_y["Energy_Available_for_Grid_POC_MW"].sum() / 4000.0, df_y["Energy_Injected_SECI_POC_MW"].sum() / 4000.0
                        exch_poc, unsold_poc = df_y["Energy_to_Exchange_POC_MW"].sum() / 4000.0, df_y["Unsold_Energy_POC_MW"].sum() / 4000.0
                        
                        s_cuf = (g_solar * 1000.0) / (solar_capacity * total_hours) * 100 if solar_capacity > 0 else 0
                        w_cuf = (g_wind * 1000.0) / (wind_capacity * total_hours) * 100 if wind_capacity > 0 else 0
                        p_cuf = ((seci_poc) / (connectivity) * 100) / 8.76
                        exp = (exch_poc / n_poc) * 100 if n_poc > 0 else 0
                        
                        yearly_compiled_snapshots[f"Year {year}"] = {
                            "cap_s": round(solar_capacity, 2), "cap_w": round(wind_capacity, 2), "cap_b": round(current_year_bess_cap, 2),
                            "conn_s": round(connectivity, 2), "conn_e": round(exchange_connectivity, 2),
                            "g_solar": round(g_solar, 2), "g_wind": round(g_wind, 2), "g_hybrid": round(g_solar+g_wind, 2),
                            "to_b": round(to_b, 2), "from_b": round(from_b, 2), "n_33_re": round(g_solar+g_wind-to_b, 2), "n_33_bess": round(g_solar+g_wind-to_b+from_b, 2),
                            "n_poc": round(n_poc, 2), "seci_poc": round(seci_poc, 2), "exch_poc": round(exch_poc, 2), "unsold_poc": round(unsold_poc, 2),
                            "sf": sf, "s_cuf": f"{s_cuf:.2f}%", "w_cuf": f"{w_cuf:.2f}%", "p_cuf": f"{p_cuf:.2f}%", "exp": f"{exp:.2f}%",
                            "s_max": round(df_y["Solar_Generation_MW"].max(), 2), "w_max": round(df_y["Wind_Generation_MW"].max(), 2), "h_max": round(df_y["Total_RE_Generation_MW"].max(), 2)
                        }

                        if year == 1:
                            df_year1_sample = df_y.copy()
                            month_data_map = {}
                            for m_idx, m_name in enumerate(months_list, 1):
                                m_group = df_y[df_y["Month_Num"] == m_idx]
                                m_hours = len(m_group) / 4.0 if len(m_group) > 0 else 1
                                ms_mu, mw_mu = m_group["Solar_Generation_MW"].sum() / 4000.0, m_group["Wind_Generation_MW"].sum() / 4000.0
                                m_to_b, m_from_b = m_group["Power_to_BESS_33KV_MWh"].sum() / 1000.0, m_group["Power_from_BESS_33KV_MWh"].sum() / 1000.0
                                mn_poc, m_seci = m_group["Energy_Available_for_Grid_POC_MW"].sum() / 4000.0, m_group["Energy_Injected_SECI_POC_MW"].sum() / 4000.0
                                m_exch, m_unsold = m_group["Energy_to_Exchange_POC_MW"].sum() / 4000.0, m_group["Unsold_Energy_POC_MW"].sum() / 4000.0
                                
                                month_data_map[m_name] = {
                                    "cap_s": round(solar_capacity, 2), "cap_w": round(wind_capacity, 2), "cap_b": round(current_year_bess_cap, 2),
                                    "conn_s": round(connectivity, 2), "conn_e": round(exchange_connectivity, 2),
                                    "g_solar": round(ms_mu, 2), "g_wind": round(mw_mu, 2), "g_hybrid": round(ms_mu+mw_mu, 2),
                                    "to_b": round(m_to_b, 2), "from_b": round(m_from_b, 2), "n_33_re": round(ms_mu+mw_mu-m_to_b, 2), "n_33_bess": round(ms_mu+mw_mu-m_to_b+m_from_b, 2),
                                    "n_poc": round(mn_poc, 2), "seci_poc": round(m_seci, 2), "exch_poc": round(m_exch, 2), "unsold_poc": round(m_unsold, 2),
                                    "sf": round(monthly_shortfall[m_idx] / 1000.0, 2), 
                                    "s_cuf": f"{(ms_mu * 1000.0) / (solar_capacity * m_hours) * 100 if solar_capacity > 0 else 0:.2f}%", 
                                    "w_cuf": f"{(mw_mu * 1000.0) / (wind_capacity * m_hours) * 100 if wind_capacity > 0 else 0:.2f}%", 
                                    "p_cuf": f"{((m_seci) / (connectivity) * 100) / (m_hours / 1000.0) if connectivity > 0 else 0:.2f}%", 
                                    "exp": f"{(m_exch / mn_poc) * 100 if mn_poc > 0 else 0:.2f}%",
                                    "s_max": round(m_group["Solar_Generation_MW"].max() if len(m_group)>0 else 0, 2),
                                    "w_max": round(m_group["Wind_Generation_MW"].max() if len(m_group)>0 else 0, 2),
                                    "h_max": round(m_group["Total_RE_Generation_MW"].max() if len(m_group)>0 else 0, 2)
                                }

                    template_blueprint = [
                        ("Solar Capacity", "MW", "cap_s"), ("Wind Capacity", "MW", "cap_w"), ("BESS Capacity", "MWh", "cap_b"),
                        ("SECI Connectivity", "MW", "conn_s"), ("Exchange Connectivity", "MW", "conn_e"), ("", "", "blank"),
                        ("Gross Solar Generation @ 33 kv PSS", "MU", "g_solar"), ("Gross Wind Generation @ 33 kv PSS", "MU", "g_wind"), ("Gross Hybrid Generation @ 33 kv PSS", "MU", "g_hybrid"), ("", "", "blank"),
                        ("Power to BESS from Hybrid @ 33 KV", "MU", "to_b"), ("Power from BESS @ 33 KV", "MU", "from_b"), ("", "", "blank"),
                        ("Net Generation @ 33 KV PSS (Solar + Wind)", "MU", "n_33_re"), ("Net Generation @ 33 KV PSS (Solar + Wind + BESS)", "MU", "n_33_bess"), ("", "", "blank"),
                        ("Net Generation @ POC", "MU", "n_poc"), ("Power to SECI @ POC", "MU", "seci_poc"), ("Power to Exchange @ POC", "MU", "exch_poc"), ("Unsold Power @ POC", "MU", "unsold_poc"), ("", "", "blank"),
                        ("Shortfall (Peak Hours Only)", "MU", "sf"), ("", "", "blank"),
                        ("Solar CuF", "%", "s_cuf"), ("Wind CuF", "%", "w_cuf"), ("CuF", "%", "p_cuf"), ("Exchange Exposure", "%", "exp"), ("", "", "blank"),
                        ("Solar Max", "MW", "s_max"), ("Wind Max", "MW", "w_max"), ("Hybrid Max", "MW", "h_max")
                    ]

                    summary_rows, monthly_rows, lifetime_rows = [], [], []
                    years_headers_list = [f"Year {y}" for y in range(1, 26)]
                    
                    for label, unit, data_key in template_blueprint:
                        summary_rows.append({"Description": label, "Unit": unit, "Generation": "" if data_key == "blank" else yearly_compiled_snapshots["Year 1"][data_key]})
                        
                        mr_dict = {"Description": label, "Unit": unit}
                        for m in months_list: mr_dict[m] = "" if data_key == "blank" else month_data_map[m][data_key]
                        mr_dict["Annual"] = "" if data_key == "blank" else yearly_compiled_snapshots["Year 1"][data_key]
                        monthly_rows.append(mr_dict)
                        
                        lr_dict = {"Description": label, "Unit": unit}
                        for yh in years_headers_list: lr_dict[yh] = "" if data_key == "blank" else yearly_compiled_snapshots[yh][data_key]
                        lifetime_rows.append(lr_dict)
                        
                    df_summary_sheet, df_monthly_matrix, df_25yr_matrix = pd.DataFrame(summary_rows), pd.DataFrame(monthly_rows), pd.DataFrame(lifetime_rows)

                    st.markdown(f"# 📊 NHPC 25-Year Lifetime Analytical Framework")
                    st.markdown(f"**Active Dual P-Factor Calibration:** Solar `{solar_p}` | Wind `{wind_p}`")
                    tab_select = st.radio("Select Active Screen View Mode", options=["📆 25-Year Lifetime Report", "📅 Baseline Monthly Report", "📋 Year-1 Summary View"], horizontal=True, key="m3_tabs")
                    
                    if tab_select == "📆 25-Year Lifetime Report": st.dataframe(df_25yr_matrix, use_container_width=True, height=750)
                    elif tab_select == "📅 Baseline Monthly Report": st.dataframe(df_monthly_matrix, use_container_width=True, height=550)
                    else: st.dataframe(df_summary_sheet, use_container_width=True, height=550)

                    df_analytical_export = df_year1_sample[["Time", "Solar_Generation_MW", "Wind_Generation_MW", "Total_RE_Generation_MW", "Energy_Injected_SECI_POC_MW", "Energy_to_Exchange_POC_MW", "Unsold_Energy_POC_MW", "Shortfall_MWh"]].copy()
                    df_analytical_export[df_analytical_export.select_dtypes(include=[np.number]).columns] = df_analytical_export.select_dtypes(include=[np.number]).round(2)

                    st.markdown("---")
                    st.markdown("### 📥 Download Unified Master Report Workbook")
                    excel_buffer = io.BytesIO()
                    with pd.ExcelWriter(excel_buffer, engine="openpyxl") as writer:
                        df_summary_sheet.to_excel(writer, index=False, sheet_name="Summary Report")
                        df_monthly_matrix.to_excel(writer, index=False, sheet_name="Monthly Cross-Tab")
                        df_25yr_matrix.to_excel(writer, index=False, sheet_name="25-Year Yearly Report") 
                        df_analytical_export.to_excel(writer, index=False, sheet_name="Analytical Data")
                    
                    st.download_button(
                        label=f"🟢 Download Unified 4-Sheet Master Report (S_{solar_p}_W_{wind_p})",
                        data=excel_buffer.getvalue(),
                        file_name=f"NHPC_BESS_Unified_25Year_Workbook_S_{solar_p}_W_{wind_p}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="m3_dl"
                    )
                else:
                    st.error("Column letters are outside sheet boundaries.")
            except Exception as e:
                st.error(f"Error processing 25-Year simulation lifecycle: {e}")
    else:
        st.info("👆 Please upload the raw data Excel file in the sidebar to proceed.")


# =========================================================================================
# MAIN APP EXECUTION & NAVIGATION
# =========================================================================================
st.sidebar.title("🎛️ Navigation Menu")
st.sidebar.markdown("Select your required analytical model below:")

app_mode = st.sidebar.radio(
    "Choose Analytical Sublayer", 
    options=[
        "Model 1: PPA RTC (Seasonal Compliance)", 
        "Model 2: PPA RTC (Standard DFR)", 
        "Model 3: NHPC BESS 25-Year Hybrid"
    ]
)

st.sidebar.markdown("---")
st.sidebar.info("This is a unified environment. Switching models dynamically changes the interface and execution pipeline entirely without interference.")

# Route traffic seamlessly without conflicts
if app_mode == "Model 1: PPA RTC (Seasonal Compliance)":
    run_model_1_rtc_v3()
elif app_mode == "Model 2: PPA RTC (Standard DFR)":
    run_model_2_rtc_v2()
elif app_mode == "Model 3: NHPC BESS 25-Year Hybrid":
    run_model_3_nhpc_25yr()
